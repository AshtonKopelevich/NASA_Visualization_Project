from openpyxl import load_workbook

def pm_pop_weighted_by_country(xlsx_path, sheet_name, scale_factor=1.0):
    """
    Returns: dict[str, list[float|None]]
      e.g. {"Greenland": [val2003, ..., val2018], "Canada": [...], ...}

    scale_factor:
      multiply each value by this (leave 1.0 to keep original sheet units;
      use 1e9 to convert kg/m^3 -> µg/m^3 if your sheet is kg/m^3)
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)

    year_to_idx = {}
    for i, v in enumerate(header):
        try:
            y = int(v)
            if 2003 <= y <= 2018:
                year_to_idx[y] = i
        except (TypeError, ValueError):
            pass

    if not year_to_idx:
        wb.close()
        raise ValueError("Could not find year columns 2003–2018 in the sheet header.")

    year_idxs = [year_to_idx[y] for y in range(2003, 2019)]

    country_col = None
    for i, v in enumerate(header):
        if isinstance(v, str) and v.strip().lower() in {
            "country", "country name", "name", "country/region"
        }:
            country_col = i
            break
    if country_col is None:
        country_col = 2

    out = {}
    for r in rows:
        if not r or len(r) <= max(year_idxs + [country_col]):
            continue

        country = r[country_col]
        if not country or not isinstance(country, str):
            continue

        values = []
        for idx in year_idxs:
            val = r[idx]
            if val is None:
                values.append(None)
            else:
                values.append(float(val) * float(scale_factor))

        out[country] = values

    wb.close()
    return out


def country_series_by_country(xlsx_path, sheet_name, divide_by=1.0):
    """
    Returns: dict[str, list[float|None]]
      e.g. {"Greenland": [val2003, ..., val2018], "Canada": [...], ...}

    divide_by:
      divide each value by this (use 10_000 for population "per 10,000 people")
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)

    year_to_idx = {}
    for i, v in enumerate(header):
        try:
            y = int(v)
            if 2003 <= y <= 2018:
                year_to_idx[y] = i
        except (TypeError, ValueError):
            pass

    if not year_to_idx:
        wb.close()
        raise ValueError("Could not find year columns 2003–2018 in the sheet header.")

    year_idxs = [year_to_idx[y] for y in range(2003, 2019)]

    country_col = None
    for i, v in enumerate(header):
        if isinstance(v, str) and v.strip().lower() in {
            "country", "country name", "name", "country/region"
        }:
            country_col = i
            break
    if country_col is None:
        country_col = 2

    out = {}
    for r in rows:
        if not r or len(r) <= max(year_idxs + [country_col]):
            continue

        country = r[country_col]
        if not country or not isinstance(country, str):
            continue

        values = []
        for idx in year_idxs:
            val = r[idx]
            if val is None:
                values.append(None)
            else:
                values.append(float(val) / float(divide_by))

        out[country] = values

    wb.close()
    return out


# Example usage:
# pollution = pm_pop_weighted_by_country("file.xlsx", "PM Pop.-Weighted (kg m^-3)", scale_factor=1.0)
# pop = country_series_by_country("file.xlsx", "Population (GPWv4.11)", divide_by=10_000)


